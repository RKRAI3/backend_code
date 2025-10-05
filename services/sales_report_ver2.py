
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Your data
data = {
    '26-09-2025': {
        'Application Support Service': {'quantity': 10.0, 'revenue': 90910.0},
        'Form Filling': {'quantity': 8.0, 'revenue': 2400.0},
        'Full Package': {'quantity': 2.0, 'revenue': 2000.0},
        'Photocopy': {'quantity': 9.0, 'revenue': 45.0},
        'Photograph': {'quantity': 11.0, 'revenue': 1100.0}
    },
    '29-09-2025': {
        'Application Support Service': {'quantity': 1.0, 'revenue': 9091.0},
        'Form Filling': {'quantity': 5.0, 'revenue': 1500.0},
        'Full Package': {'quantity': 1.0, 'revenue': 1000.0},
        'Photocopy': {'quantity': 10.0, 'revenue': 50.0},
        'Photograph': {'quantity': 5.0, 'revenue': 500.0}
    }
}

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales Report"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

# Get all unique services
all_services = set()
for date_data in data.values():
    all_services.update(date_data.keys())
services_list = sorted(all_services)

# Create headers
ws['A1'] = 'Date'
ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = center_align
ws['A1'].border = border

col_idx = 2
for service in services_list:
    # Service name header (spans 2 columns)
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
    cell = ws.cell(row=1, column=col_idx)
    cell.value = service
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    ws.cell(row=1, column=col_idx+1).border = border
    
    # Quantity and Revenue sub-headers
    qty_cell = ws.cell(row=2, column=col_idx)
    qty_cell.value = 'Quantity'
    qty_cell.fill = subheader_fill
    qty_cell.font = subheader_font
    qty_cell.alignment = center_align
    qty_cell.border = border
    
    rev_cell = ws.cell(row=2, column=col_idx+1)
    rev_cell.value = 'Revenue'
    rev_cell.fill = subheader_fill
    rev_cell.font = subheader_font
    rev_cell.alignment = center_align
    rev_cell.border = border
    
    col_idx += 2

# Merge Date header cell
ws.merge_cells('A1:A2')

# Write data
row = 3
for date, services in data.items():
    ws.cell(row=row, column=1).value = date
    ws.cell(row=row, column=1).alignment = center_align
    ws.cell(row=row, column=1).border = border
    
    col_idx = 2
    for service in services_list:
        if service in services:
            # Quantity
            qty_cell = ws.cell(row=row, column=col_idx)
            qty_cell.value = services[service]['quantity']
            qty_cell.alignment = right_align
            qty_cell.border = border
            qty_cell.number_format = '0.0'
            
            # Revenue
            rev_cell = ws.cell(row=row, column=col_idx+1)
            rev_cell.value = services[service]['revenue']
            rev_cell.alignment = right_align
            rev_cell.border = border
            rev_cell.number_format = '₹#,##0.00'
        else:
            # Empty cells with borders
            ws.cell(row=row, column=col_idx).border = border
            ws.cell(row=row, column=col_idx+1).border = border
        
        col_idx += 2
    
    row += 1

# Add totals row
ws.cell(row=row, column=1).value = 'TOTAL'
ws.cell(row=row, column=1).fill = subheader_fill
ws.cell(row=row, column=1).font = subheader_font
ws.cell(row=row, column=1).alignment = center_align
ws.cell(row=row, column=1).border = border

col_idx = 2
for service in services_list:
    # Total Quantity
    qty_cell = ws.cell(row=row, column=col_idx)
    qty_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx)}3:{openpyxl.utils.get_column_letter(col_idx)}{row-1})'
    qty_cell.fill = subheader_fill
    qty_cell.font = subheader_font
    qty_cell.alignment = right_align
    qty_cell.border = border
    qty_cell.number_format = '0.0'
    
    # Total Revenue
    rev_cell = ws.cell(row=row, column=col_idx+1)
    rev_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx+1)}3:{openpyxl.utils.get_column_letter(col_idx+1)}{row-1})'
    rev_cell.fill = subheader_fill
    rev_cell.font = subheader_font
    rev_cell.alignment = right_align
    rev_cell.border = border
    rev_cell.number_format = '₹#,##0.00'
    
    col_idx += 2

# Adjust column widths
ws.column_dimensions['A'].width = 15
for i in range(2, col_idx):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

# Save the file
filename = 'sales_report_v2.xlsx'
wb.save(filename)
print(f"Excel file '{filename}' created successfully!")