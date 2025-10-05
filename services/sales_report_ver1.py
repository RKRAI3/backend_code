# -*- coding: utf-8 -*-
"""
Created on Wed Oct  1 12:37:47 2025

@author: RAVI KANT
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Your data
data = {
    '26-09-2025': {
        'Application Support Service': {'quantity': 10.0, 'revenue': 90910.0},
        'Form Filling': {'quantity': 8.0, 'revenue': 2400.0},
        'Full Package': {'quantity': 2.0, 'revenue': 2000.0},
        'Photocopy': {'quantity': 9.0, 'revenue': 45.0},
        'Photograph': {'quantity': 11.0, 'revenue': 1100.0}
    },
    '29-09-2025': {
        'Application Support Service': {'quantity': 1.0, 'revenue': 9091.0},
        'Form Filling': {'quantity': 5.0, 'revenue': 1500.0},
        'Full Package': {'quantity': 1.0, 'revenue': 1000.0},
        'Photocopy': {'quantity': 10.0, 'revenue': 50.0},
        'Photograph': {'quantity': 5.0, 'revenue': 500.0}
    }
}

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales Report"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

# Write headers
ws['A1'] = 'Date'
ws['B1'] = 'Service'
ws['C1'] = 'Quantity'
ws['D1'] = 'Revenue'

# Apply header styling
for col in ['A', 'B', 'C', 'D']:
    cell = ws[f'{col}1']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Write data
row = 2
for date, services in data.items():
    date_start_row = row
    
    for service_name, values in services.items():
        ws[f'A{row}'] = date
        ws[f'B{row}'] = service_name
        ws[f'C{row}'] = values['quantity']
        ws[f'D{row}'] = values['revenue']
        
        # Apply borders
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = border
        
        # Align numbers to the right
        ws[f'C{row}'].alignment = right_align
        ws[f'D{row}'].alignment = right_align
        
        row += 1
    
    # Merge date cells for better readability
    if row - date_start_row > 1:
        ws.merge_cells(f'A{date_start_row}:A{row-1}')
        ws[f'A{date_start_row}'].alignment = center_align

# Add totals row
ws[f'A{row}'] = 'TOTAL'
ws[f'B{row}'] = ''
ws[f'C{row}'] = f'=SUM(C2:C{row-1})'
ws[f'D{row}'] = f'=SUM(D2:D{row-1})'

# Style totals row
for col in ['A', 'B', 'C', 'D']:
    cell = ws[f'{col}{row}']
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border
    cell.alignment = center_align if col == 'A' else right_align

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15

# Number formatting
for r in range(2, row + 1):
    ws[f'D{r}'].number_format = '₹#,##0.00'
    ws[f'C{r}'].number_format = '0.0'

# Save the file
filename = 'sales_report.xlsx'
wb.save(filename)
print(f"Excel file '{filename}' created successfully!")