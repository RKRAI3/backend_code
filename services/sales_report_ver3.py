from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import jsonify


from openpyxl.utils import get_column_letter

def add_quantity_and_revenue_sheets(wb, dashboard_data, currency='₿'):
    quantity_report = dashboard_data.get("quantity_report", [])
    revenue_report = dashboard_data.get("revenue_report", [])

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True)

    # Helper function to write a full sheet (used for both reports)
    def write_report(sheet_name, data, is_currency=False):
        if not data:
            return

        ws = wb.create_sheet(sheet_name)
        headers = list(data[0].keys())

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Write data rows
        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = record[header]
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

                if isinstance(value, (int, float)):
                    if is_currency and header not in ["receipt_number", "created_at"]:
                        cell.number_format = f'"{currency}"#,##0.00'

        # Add totals row at the end (sum of all numeric columns)
        last_row = len(data) + 2
        ws.cell(row=last_row, column=1, value="TOTALS").font = bold_font
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='right')

        for col_num, header in enumerate(headers, 1):
            if header not in ["receipt_number", "created_at"]:
                col_letter = get_column_letter(col_num)
                formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
                total_cell = ws.cell(row=last_row, column=col_num, value=formula)
                total_cell.font = bold_font
                total_cell.border = thin_border
                if is_currency:
                    total_cell.number_format = f'"{currency}"#,##0.00'

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

    # Create both sheets
    write_report("Quantity Report", quantity_report, is_currency=False)
    write_report("Revenue Report", revenue_report, is_currency=True)


def generate_comprehensive_excel_report(dashboard_data, currency):
    """
    Generate comprehensive Excel report with multiple sheets
    
    Args:
        dashboard_data: Dictionary containing all dashboard data
        
    Returns:
        dict with 'blob' (base64 string), 'filename', and 'mimeType'
    """
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define common styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ==================== SHEET 1: Summary ====================
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Title
    ws_summary['A1'] = 'Dashboard Summary Report'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')
    
    # Data
    summary_data = [
        ['Period', dashboard_data.get('period', 'N/A')],
        ['Growth Trend', dashboard_data.get('growthTrend', 'N/A')],
        ['Total Products', dashboard_data.get('totalProducts', 0)],
        ['Unique Products', dashboard_data.get('uniqueProducts', 0)],
        ['Total Receipts', dashboard_data.get('totalReceipts', 0)],
        ['Total Revenue (₿)', dashboard_data.get('totalRevenue', 0)],
        ['Standard Revenue (₿)', dashboard_data.get('totalStdRevenue', 0)],
        ['Revenue Trend', dashboard_data.get('revenueTrend', 'N/A')],
        ['Receipts Trend', dashboard_data.get('receiptsTrend', 'N/A')],
        ['Product Trend', dashboard_data.get('productTrend', 'N/A')],
        ['Unique Product Trend', dashboard_data.get('uniqueProductTrend', 'N/A')]
    ]
    
    row = 3
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'A{row}'].fill = subheader_fill
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        ws_summary[f'B{row}'].alignment = right_align
        row += 1
    
    # Format currency cells '₿#,##0'
    for r in [8, 9]:
        ws_summary[f'B{r}'].number_format = f'{currency}#,##0'
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # ==================== SHEET 2: Product Revenue ====================
    ws_product = wb.create_sheet("Product Revenue", 1)
    
    # Headers
    headers = ['Product Name', 'Quantity Sold', 'Revenue (₿)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_product.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    product_revenue = dashboard_data.get('productRevenue', [])
    for row_idx, product in enumerate(product_revenue, start=2):
        ws_product.cell(row=row_idx, column=1, value=product.get('name', ''))
        ws_product.cell(row=row_idx, column=2, value=product.get('quantity', 0))
        ws_product.cell(row=row_idx, column=3, value=product.get('revenue', 0))
        
        for col_idx in range(1, 4):
            cell = ws_product.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx > 1:
                cell.alignment = right_align
        
        # Format numbers
        ws_product.cell(row=row_idx, column=3).number_format = f'{currency}#,##0'
    
    # Totals row
    total_row = len(product_revenue) + 2
    ws_product.cell(row=total_row, column=1, value='TOTAL')
    ws_product.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row-1})')
    ws_product.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})')
    
    for col_idx in range(1, 4):
        cell = ws_product.cell(row=total_row, column=col_idx)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.border = border
        if col_idx > 1:
            cell.alignment = right_align
    
    ws_product.cell(row=total_row, column=3).number_format = f'{currency}#,##0'
    
    ws_product.column_dimensions['A'].width = 28
    ws_product.column_dimensions['B'].width = 15
    ws_product.column_dimensions['C'].width = 15
    
    # ==================== SHEET 3: Price Comparison ====================
    ws_price = wb.create_sheet("Price Comparison", 2)
    
    # Headers
    headers = ['Product Name', 'Standard Price (₿)', 'Vendor Price (₿)', 'Difference (₿)', 'Variance %']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_price.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    price_comparison = dashboard_data.get('priceComparison', [])
    for row_idx, product in enumerate(price_comparison, start=2):
        std_price = product.get('std_price', 0)
        vend_price = product.get('vend_price', 0)
        difference = std_price - vend_price
        variance = ((std_price - vend_price) / std_price * 100) if std_price != 0 else 0
        
        ws_price.cell(row=row_idx, column=1, value=product.get('name', ''))
        ws_price.cell(row=row_idx, column=2, value=std_price)
        ws_price.cell(row=row_idx, column=3, value=vend_price)
        ws_price.cell(row=row_idx, column=4, value=difference)
        ws_price.cell(row=row_idx, column=5, value=variance)
        
        for col_idx in range(1, 6):
            cell = ws_price.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx > 1:
                cell.alignment = right_align
        
        # Format currency
        for col in [2, 3, 4]:
            ws_price.cell(row=row_idx, column=col).number_format = '₿#,##0.00'
        ws_price.cell(row=row_idx, column=5).number_format = '0.00"%"'
        
        # Color code based on variance
        variance_cell = ws_price.cell(row=row_idx, column=5)
        if variance > 20:
            variance_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif variance < 0:
            variance_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    ws_price.column_dimensions['A'].width = 28
    ws_price.column_dimensions['B'].width = 18
    ws_price.column_dimensions['C'].width = 18
    ws_price.column_dimensions['D'].width = 15
    ws_price.column_dimensions['E'].width = 12
    
    # ==================== SHEET 4: Sales Revenue Data ====================
    # Generate the detailed sales report from previous code
    ws_sales = wb.create_sheet("Sales Revenue Data", 3)
    
    # Get sales data grouped by date
    quantity_report = dashboard_data.get('quantity_report', [])
    revenue_report = dashboard_data.get('revenue_report', [])
    
    # Group data by date
    sales_by_date = {}
    for qty_entry, rev_entry in zip(quantity_report, revenue_report):
        date = qty_entry.get('created_at')
        if date not in sales_by_date:
            sales_by_date[date] = {}
        
        for key, value in qty_entry.items():
            if key not in ['created_at', 'receipt_number']:
                if key not in sales_by_date[date]:
                    sales_by_date[date][key] = {'quantity': 0, 'revenue': 0}
                sales_by_date[date][key]['quantity'] += value
                sales_by_date[date][key]['revenue'] += rev_entry.get(key, 0)
    
    # Get all unique services
    all_services = set()
    for date_data in sales_by_date.values():
        all_services.update(date_data.keys())
    services_list = sorted(all_services)
    
    # Create headers
    ws_sales['A1'] = 'Date'
    ws_sales['A1'].fill = header_fill
    ws_sales['A1'].font = header_font
    ws_sales['A1'].alignment = center_align
    ws_sales['A1'].border = border
    
    col_idx = 2
    for service in services_list:
        # Service name header (spans 2 columns)
        ws_sales.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
        cell = ws_sales.cell(row=1, column=col_idx)
        cell.value = service
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws_sales.cell(row=1, column=col_idx+1).border = border
        
        # Quantity and Revenue sub-headers
        qty_cell = ws_sales.cell(row=2, column=col_idx)
        qty_cell.value = 'Quantity'
        qty_cell.fill = subheader_fill
        qty_cell.font = subheader_font
        qty_cell.alignment = center_align
        qty_cell.border = border
        
        rev_cell = ws_sales.cell(row=2, column=col_idx+1)
        rev_cell.value = 'Revenue'
        rev_cell.fill = subheader_fill
        rev_cell.font = subheader_font
        rev_cell.alignment = center_align
        rev_cell.border = border
        
        col_idx += 2
    
    # Add Total Revenue column
    total_col = col_idx
    ws_sales.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)
    total_header = ws_sales.cell(row=1, column=total_col)
    total_header.value = 'Total Revenue'
    total_header.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    total_header.font = Font(bold=True, color="FFFFFF", size=11)
    total_header.alignment = center_align
    total_header.border = border
    
    # Merge Date header cell
    ws_sales.merge_cells('A1:A2')
    
    # Write data
    row = 3
    for date, services in sorted(sales_by_date.items()):
        ws_sales.cell(row=row, column=1, value=date)
        ws_sales.cell(row=row, column=1).alignment = center_align
        ws_sales.cell(row=row, column=1).border = border
        
        col_idx = 2
        for service in services_list:
            if service in services:
                # Quantity
                qty_cell = ws_sales.cell(row=row, column=col_idx)
                qty_cell.value = services[service]['quantity']
                qty_cell.alignment = right_align
                qty_cell.border = border
                qty_cell.number_format = '0'
                
                # Revenue
                rev_cell = ws_sales.cell(row=row, column=col_idx+1)
                rev_cell.value = services[service]['revenue']
                rev_cell.alignment = right_align
                rev_cell.border = border
                rev_cell.number_format = f'{currency}#,##0'
            else:
                ws_sales.cell(row=row, column=col_idx).border = border
                ws_sales.cell(row=row, column=col_idx+1).border = border
            
            col_idx += 2
        
        # Add Total Revenue for this date
        total_cell = ws_sales.cell(row=row, column=total_col)
        revenue_cols = [openpyxl.utils.get_column_letter(c) for c in range(3, total_col, 2)]
        total_formula = '+'.join([f'{col}{row}' for col in revenue_cols])
        total_cell.value = f'={total_formula}'
        total_cell.alignment = right_align
        total_cell.border = border
        total_cell.number_format = f'{currency}#,##0'
        total_cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
        row += 1
    
    # Add totals row
    ws_sales.cell(row=row, column=1, value='TOTAL')
    ws_sales.cell(row=row, column=1).fill = subheader_fill
    ws_sales.cell(row=row, column=1).font = subheader_font
    ws_sales.cell(row=row, column=1).alignment = center_align
    ws_sales.cell(row=row, column=1).border = border
    
    col_idx = 2
    for service in services_list:
        # Total Quantity
        qty_cell = ws_sales.cell(row=row, column=col_idx)
        qty_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx)}3:{openpyxl.utils.get_column_letter(col_idx)}{row-1})'
        qty_cell.fill = subheader_fill
        qty_cell.font = subheader_font
        qty_cell.alignment = right_align
        qty_cell.border = border
        qty_cell.number_format = '0'
        
        # Total Revenue
        rev_cell = ws_sales.cell(row=row, column=col_idx+1)
        rev_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx+1)}3:{openpyxl.utils.get_column_letter(col_idx+1)}{row-1})'
        rev_cell.fill = subheader_fill
        rev_cell.font = subheader_font
        rev_cell.alignment = right_align
        rev_cell.border = border
        rev_cell.number_format = f'{currency}#,##0'
        
        col_idx += 2
    
    # Grand total
    grand_total_cell = ws_sales.cell(row=row, column=total_col)
    grand_total_cell.value = f'=SUM({openpyxl.utils.get_column_letter(total_col)}3:{openpyxl.utils.get_column_letter(total_col)}{row-1})'
    grand_total_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    grand_total_cell.font = Font(bold=True, color="FFFFFF", size=10)
    grand_total_cell.alignment = right_align
    grand_total_cell.border = border
    grand_total_cell.number_format = f'{currency}#,##0'
    
    # Adjust column widths
    ws_sales.column_dimensions['A'].width = 15
    for i in range(2, total_col + 1):
        ws_sales.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    
    add_quantity_and_revenue_sheets(wb, dashboard_data, currency='₿')
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    
    # Convert to base64
    excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
    
    return {
        'blob': excel_base64,
        'filename': 'dashboard_comprehensive_report.xlsx',
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
# from io import BytesIO
# import base64
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from flask import jsonify

# def generate_excel_blob(data):
#     """
#     Generate Excel file and return as base64 encoded blob
    
#     Args:
#         data: Dictionary with date as keys and services as nested dict
        
#     Returns:
#         dict with 'blob' (base64 string) and 'filename'
#     """
#     # Create workbook in memory
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Sales Report"
    
#     # Define styles
#     header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#     header_font = Font(bold=True, color="FFFFFF", size=11)
#     subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
#     subheader_font = Font(bold=True, size=10)
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
#     center_align = Alignment(horizontal='center', vertical='center')
#     right_align = Alignment(horizontal='right', vertical='center')
    
#     # Get all unique services
#     all_services = set()
#     for date_data in data.values():
#         all_services.update(date_data.keys())
#     services_list = sorted(all_services)
    
#     # Create headers
#     ws['A1'] = 'Date'
#     ws['A1'].fill = header_fill
#     ws['A1'].font = header_font
#     ws['A1'].alignment = center_align
#     ws['A1'].border = border
    
#     col_idx = 2
#     for service in services_list:
#         # Service name header (spans 2 columns)
#         ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
#         cell = ws.cell(row=1, column=col_idx)
#         cell.value = service
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center_align
#         cell.border = border
#         ws.cell(row=1, column=col_idx+1).border = border
        
#         # Quantity and Revenue sub-headers
#         qty_cell = ws.cell(row=2, column=col_idx)
#         qty_cell.value = 'Quantity'
#         qty_cell.fill = subheader_fill
#         qty_cell.font = subheader_font
#         qty_cell.alignment = center_align
#         qty_cell.border = border
        
#         rev_cell = ws.cell(row=2, column=col_idx+1)
#         rev_cell.value = 'Revenue'
#         rev_cell.fill = subheader_fill
#         rev_cell.font = subheader_font
#         rev_cell.alignment = center_align
#         rev_cell.border = border
        
#         col_idx += 2
    
#     # Add Total Revenue column header
#     total_col = col_idx
#     ws.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)
#     total_header = ws.cell(row=1, column=total_col)
#     total_header.value = 'Total Revenue'
#     total_header.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
#     total_header.font = Font(bold=True, color="FFFFFF", size=11)
#     total_header.alignment = center_align
#     total_header.border = border
    
#     # Merge Date header cell
#     ws.merge_cells('A1:A2')
    
#     # Write data
#     row = 3
#     for date, services in data.items():
#         ws.cell(row=row, column=1).value = date
#         ws.cell(row=row, column=1).alignment = center_align
#         ws.cell(row=row, column=1).border = border
        
#         col_idx = 2
#         for service in services_list:
#             if service in services:
#                 # Quantity
#                 qty_cell = ws.cell(row=row, column=col_idx)
#                 qty_cell.value = services[service]['quantity']
#                 qty_cell.alignment = right_align
#                 qty_cell.border = border
#                 qty_cell.number_format = '0.0'
                
#                 # Revenue
#                 rev_cell = ws.cell(row=row, column=col_idx+1)
#                 rev_cell.value = services[service]['revenue']
#                 rev_cell.alignment = right_align
#                 rev_cell.border = border
#                 rev_cell.number_format = '₿#,##0.00'
#             else:
#                 # Empty cells with borders
#                 ws.cell(row=row, column=col_idx).border = border
#                 ws.cell(row=row, column=col_idx+1).border = border
            
#             col_idx += 2
        
#         # Add Total Revenue for this date
#         total_cell = ws.cell(row=row, column=total_col)
#         revenue_cols = [openpyxl.utils.get_column_letter(c) for c in range(3, total_col, 2)]
#         total_formula = '+'.join([f'{col}{row}' for col in revenue_cols])
#         total_cell.value = f'={total_formula}'
#         total_cell.alignment = right_align
#         total_cell.border = border
#         total_cell.number_format = '₿#,##0.00'
#         total_cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
#         row += 1
    
#     # Add totals row
#     ws.cell(row=row, column=1).value = 'TOTAL'
#     ws.cell(row=row, column=1).fill = subheader_fill
#     ws.cell(row=row, column=1).font = subheader_font
#     ws.cell(row=row, column=1).alignment = center_align
#     ws.cell(row=row, column=1).border = border
    
#     col_idx = 2
#     for service in services_list:
#         # Total Quantity
#         qty_cell = ws.cell(row=row, column=col_idx)
#         qty_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx)}3:{openpyxl.utils.get_column_letter(col_idx)}{row-1})'
#         qty_cell.fill = subheader_fill
#         qty_cell.font = subheader_font
#         qty_cell.alignment = right_align
#         qty_cell.border = border
#         qty_cell.number_format = '0.0'
        
#         # Total Revenue
#         rev_cell = ws.cell(row=row, column=col_idx+1)
#         rev_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx+1)}3:{openpyxl.utils.get_column_letter(col_idx+1)}{row-1})'
#         rev_cell.fill = subheader_fill
#         rev_cell.font = subheader_font
#         rev_cell.alignment = right_align
#         rev_cell.border = border
#         rev_cell.number_format = '₿#,##0.00'
        
#         col_idx += 2
    
#     # Add grand total for Total Revenue column
#     grand_total_cell = ws.cell(row=row, column=total_col)
#     grand_total_cell.value = f'=SUM({openpyxl.utils.get_column_letter(total_col)}3:{openpyxl.utils.get_column_letter(total_col)}{row-1})'
#     grand_total_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
#     grand_total_cell.font = Font(bold=True, color="FFFFFF", size=10)
#     grand_total_cell.alignment = right_align
#     grand_total_cell.border = border
#     grand_total_cell.number_format = '₿#,##0.00'
    
#     # Adjust column widths
#     ws.column_dimensions['A'].width = 15
#     for i in range(2, col_idx):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
#     ws.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 16
    
#     # Save to BytesIO object (in-memory)
#     # IMPORTANT: Use BytesIO to preserve all formatting
#     excel_buffer = BytesIO()
    
#     # Save with full formatting support
#     wb.save(excel_buffer)
    
#     # Get the bytes
#     excel_bytes = excel_buffer.getvalue()
    
#     # Convert to base64
#     excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
    
#     return {
#         'blob': excel_base64,
#         'filename': 'sales_report.xlsx',
#         'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     }

# # Create workbook and worksheet
# def generate_excel_blob(data):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Sales Report"

#     # Define styles
#     header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#     header_font = Font(bold=True, color="FFFFFF", size=11)
#     subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
#     subheader_font = Font(bold=True, size=10)
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
#     center_align = Alignment(horizontal='center', vertical='center')
#     right_align = Alignment(horizontal='right', vertical='center')

#     # Get all unique services
#     all_services = set()
#     for date_data in data.values():
#         all_services.update(date_data.keys())
#     services_list = sorted(all_services)

#     # Create headers
#     ws['A1'] = 'Date'
#     ws['A1'].fill = header_fill
#     ws['A1'].font = header_font
#     ws['A1'].alignment = center_align
#     ws['A1'].border = border

#     col_idx = 2
#     for service in services_list:
#         # Service name header (spans 2 columns)
#         ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
#         cell = ws.cell(row=1, column=col_idx)
#         cell.value = service
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center_align
#         cell.border = border
#         ws.cell(row=1, column=col_idx+1).border = border
        
#         # Quantity and Revenue sub-headers
#         qty_cell = ws.cell(row=2, column=col_idx)
#         qty_cell.value = 'Quantity'
#         qty_cell.fill = subheader_fill
#         qty_cell.font = subheader_font
#         qty_cell.alignment = center_align
#         qty_cell.border = border
        
#         rev_cell = ws.cell(row=2, column=col_idx+1)
#         rev_cell.value = 'Revenue'
#         rev_cell.fill = subheader_fill
#         rev_cell.font = subheader_font
#         rev_cell.alignment = center_align
#         rev_cell.border = border
        
#         col_idx += 2

#     # Merge Date header cell
#     ws.merge_cells('A1:A2')

#     # Add Total Revenue column header
#     total_col = col_idx
#     ws.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)
#     total_header = ws.cell(row=1, column=total_col)
#     total_header.value = 'Total Revenue'
#     total_header.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
#     total_header.font = Font(bold=True, color="FFFFFF", size=11)
#     total_header.alignment = center_align
#     total_header.border = border

#     # Write data
#     row = 3
#     for date, services in data.items():
#         ws.cell(row=row, column=1).value = date
#         ws.cell(row=row, column=1).alignment = center_align
#         ws.cell(row=row, column=1).border = border
        
#         col_idx = 2
#         for service in services_list:
#             if service in services:
#                 # Quantity
#                 qty_cell = ws.cell(row=row, column=col_idx)
#                 qty_cell.value = services[service]['quantity']
#                 qty_cell.alignment = right_align
#                 qty_cell.border = border
#                 qty_cell.number_format = '0.0'
                
#                 # Revenue
#                 rev_cell = ws.cell(row=row, column=col_idx+1)
#                 rev_cell.value = services[service]['revenue']
#                 rev_cell.alignment = right_align
#                 rev_cell.border = border
#                 rev_cell.number_format = '₿#,##0.00'
#             else:
#                 # Empty cells with borders
#                 ws.cell(row=row, column=col_idx).border = border
#                 ws.cell(row=row, column=col_idx+1).border = border
            
#             col_idx += 2
        
#         # Add Total Revenue for this date
#         total_cell = ws.cell(row=row, column=total_col)
#         revenue_cols = [openpyxl.utils.get_column_letter(c) for c in range(3, total_col, 2)]
#         total_formula = '+'.join([f'{col}{row}' for col in revenue_cols])
#         total_cell.value = f'={total_formula}'
#         total_cell.alignment = right_align
#         total_cell.border = border
#         total_cell.number_format = '₿#,##0.00'
#         total_cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
#         row += 1

#     # Add totals row
#     ws.cell(row=row, column=1).value = 'TOTAL'
#     ws.cell(row=row, column=1).fill = subheader_fill
#     ws.cell(row=row, column=1).font = subheader_font
#     ws.cell(row=row, column=1).alignment = center_align
#     ws.cell(row=row, column=1).border = border

#     col_idx = 2
#     for service in services_list:
#         # Total Quantity
#         qty_cell = ws.cell(row=row, column=col_idx)
#         qty_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx)}3:{openpyxl.utils.get_column_letter(col_idx)}{row-1})'
#         qty_cell.fill = subheader_fill
#         qty_cell.font = subheader_font
#         qty_cell.alignment = right_align
#         qty_cell.border = border
#         qty_cell.number_format = '0.0'
        
#         # Total Revenue
#         rev_cell = ws.cell(row=row, column=col_idx+1)
#         rev_cell.value = f'=SUM({openpyxl.utils.get_column_letter(col_idx+1)}3:{openpyxl.utils.get_column_letter(col_idx+1)}{row-1})'
#         rev_cell.fill = subheader_fill
#         rev_cell.font = subheader_font
#         rev_cell.alignment = right_align
#         rev_cell.border = border
#         rev_cell.number_format = '₿#,##0.00'
        
#         col_idx += 2

#     # Add grand total for Total Revenue column
#     grand_total_cell = ws.cell(row=row, column=total_col)
#     grand_total_cell.value = f'=SUM({openpyxl.utils.get_column_letter(total_col)}3:{openpyxl.utils.get_column_letter(total_col)}{row-1})'
#     grand_total_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
#     grand_total_cell.font = Font(bold=True, color="FFFFFF", size=10)
#     grand_total_cell.alignment = right_align
#     grand_total_cell.border = border
#     grand_total_cell.number_format = '₿#,##0.00'

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 15
#     for i in range(2, col_idx):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
#     ws.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 16
#     excel_buffer = BytesIO()
#     wb.save(excel_buffer)
#     excel_buffer.seek(0)
    
#     # Convert to base64
#     excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
    
#     return {
#         'blob': excel_base64,
#         'filename': 'sales_report.xlsx',
#         'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     }
# # Save the file
# # filename = 'sales_report_v3.xlsx'
# # wb.save(filename)
# # print(f"Excel file '{filename}' created successfully!")